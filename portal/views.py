from __future__ import annotations

import os
from io import BytesIO
from datetime import datetime, time, timedelta, date
from typing import Any, Dict, List, Optional

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from accounts.models import User
from website.site_settings import get_site_settings

from .forms import TicketCreateForm
from .forms_site_settings import SitePagesSettingsForm
from .forms_status import RequestStatusForm
from .forms_technician import TechnicianForm
from .models import (
    PublicRequest,
    PublicRequestAttachment,
    RequestStatus,
    Technician,
    Ticket,
    AbuseEvent,   # adaugi în models.py
    BlockedIP,    # adaugi în models.py
)
from .permissions import role_required


# ============================================================
# Helpers
# ============================================================

def get_client_ip(request) -> Optional[str]:
    """
    În spatele unui reverse proxy, X_FORWARDED_FOR poate conține mai multe IP-uri.
    Primul e, de regulă, clientul.
    """
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def log_abuse(request, reason: str) -> None:
    AbuseEvent.objects.create(
        ip=get_client_ip(request),
        user=request.user if request.user.is_authenticated else None,
        path=(request.path or "")[:255],
        reason=reason[:120],
        user_agent=(request.META.get("HTTP_USER_AGENT", "") or "")[:255],
    )


def maybe_block_ip(ip: Optional[str], reason: str) -> None:
    """
    Dacă avem >=30 abuzuri în ultimele 10 minute -> blocăm IP 2 ore.
    """
    if not ip:
        return
    since = timezone.now() - timedelta(minutes=10)
    cnt = AbuseEvent.objects.filter(ip=ip, created_at__gte=since).count()
    if cnt >= 30:
        BlockedIP.objects.update_or_create(
            ip=ip,
            defaults={
                "blocked_until": timezone.now() + timedelta(hours=2),
                "reason": reason[:120],
            }
        )


# ----------------------------
# DB Rate-limit (fără librării)
# ----------------------------

def rl_key_for_request(request, base: str) -> str:
    """
    Cheie stabilă per user dacă e logat, altfel per IP.
    """
    if request.user.is_authenticated:
        return f"{base}:u:{request.user.id}"
    return f"{base}:ip:{get_client_ip(request) or 'unknown'}"


def db_rate_limit_hit(request, base: str, limit: int, window: timedelta) -> bool:
    """
    True dacă a depășit limita în fereastra de timp.
    Folosește AbuseEvent ca jurnal/contor.
    """
    key = rl_key_for_request(request, base)
    since = timezone.now() - window
    return AbuseEvent.objects.filter(reason=key, created_at__gte=since).count() >= limit


def db_rate_limit_mark(request, base: str) -> None:
    """
    Marchează o tentativă (increment implicit).
    """
    key = rl_key_for_request(request, base)
    log_abuse(request, key)


def _safe_int(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _normalize(s: Optional[str]) -> str:
    return (s or "").strip()


def _normalize_lower(s: Optional[str]) -> str:
    return _normalize(s).lower()


def _parse_date(s: Optional[str]) -> Optional[date]:
    s = _normalize(s)
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _month_range(d: date) -> tuple[date, date]:
    first = d.replace(day=1)
    if first.month == 12:
        next_first = first.replace(year=first.year + 1, month=1, day=1)
    else:
        next_first = first.replace(month=first.month + 1, day=1)
    last = next_first - timedelta(days=1)
    return first, last


def _qs_keep(request, drop=None, **overrides) -> str:
    q = request.GET.copy()
    drop = drop or []
    for k in drop:
        q.pop(k, None)

    for k, v in overrides.items():
        if v is None:
            q.pop(k, None)
        else:
            q[k] = str(v)

    return q.urlencode()


def _build_autocomplete(items: List[Dict[str, Any]], key: str, limit: int = 80) -> List[str]:
    seen = set()
    out = []
    for it in items:
        val = str(it.get(key) or "").strip()
        if not val:
            continue
        if val in seen:
            continue
        seen.add(val)
        out.append(val)
        if len(out) >= limit:
            break
    return out


def _user_name(u):
    if not u:
        return "-"
    full = (f"{getattr(u, 'first_name', '')} {getattr(u, 'last_name', '')}").strip()
    return full or getattr(u, "email", "-")


def _user_phone(u):
    return getattr(u, "phone", None) or getattr(u, "phone_number", None) or "-"


def _ticket_source(t: Ticket) -> str:
    if t.created_by and getattr(t.created_by, "role", None) == User.Role.CLIENT:
        return "CLIENT"
    return "INTERN"


def _public_source(r: PublicRequest) -> str:
    if not getattr(r, "user_id", None):
        return "PUBLIC"
    u = getattr(r, "user", None)
    if u and getattr(u, "role", None) == User.Role.CLIENT:
        return "CLIENT"
    return "INTERN"


# ============================================================
# Tickets CRUD (ADMIN/MANAGER)
# ============================================================

@role_required(User.Role.ADMIN, User.Role.MANAGER)
def ticket_edit(request, pk):
    t = get_object_or_404(Ticket, pk=pk)

    if request.method == "POST":
        t.subject = request.POST.get("subject", t.subject)
        t.message = request.POST.get("message", t.message)

        status = request.POST.get("status")
        if status and status in dict(Ticket.Status.choices):
            t.status = status

        assigned_to_id = request.POST.get("assigned_to_id") or None
        t.assigned_to_id = int(assigned_to_id) if assigned_to_id else None

        t.save()
        messages.success(request, "Cererea a fost actualizată.")
        return redirect("portal_dashboard")

    return render(request, "portal/ticket_edit.html", {
        "ticket": t,
        "statuses": Ticket.Status.choices,
        "tech_users": User.objects.filter(role=User.Role.TEHNICIAN, is_active=True),
    })


@role_required(User.Role.ADMIN, User.Role.MANAGER)
@require_POST
def ticket_delete(request, pk):
    t = get_object_or_404(Ticket, pk=pk)
    t.delete()
    messages.success(request, "Cererea a fost ștearsă.")
    return redirect("portal_dashboard")


# ============================================================
# PublicRequest edit/delete + attachments + assigned technician
# ============================================================

@role_required(User.Role.ADMIN, User.Role.MANAGER)
def public_request_edit(request, pk):
    r = get_object_or_404(
        PublicRequest.objects.select_related("user", "status", "assigned_to").prefetch_related("attachments"),
        pk=pk
    )

    if request.method == "POST":
        r.email = request.POST.get("email", r.email)
        r.phone = request.POST.get("phone", r.phone)
        r.company = request.POST.get("company", r.company)
        r.company_cif = request.POST.get("company_cif", r.company_cif)
        r.description = request.POST.get("description", r.description)

        status_id = request.POST.get("status_id") or None
        r.status_id = int(status_id) if status_id else None

        assigned_to_id = request.POST.get("assigned_to_id") or None
        r.assigned_to_id = int(assigned_to_id) if assigned_to_id else None

        r.save()

        files = request.FILES.getlist("attachments")
        if files:
            allowed_ext = {".pdf", ".png", ".jpg", ".jpeg", ".doc", ".docx", ".xls", ".xlsx", ".txt"}
            max_mb = 25

            for f in files:
                ext = os.path.splitext(f.name.lower())[1]
                if ext and ext not in allowed_ext:
                    messages.error(request, f"Fișierul {f.name} are extensie neacceptată.")
                    continue
                if f.size > max_mb * 1024 * 1024:
                    messages.error(request, f"Fișierul {f.name} depășește {max_mb}MB.")
                    continue

                PublicRequestAttachment.objects.create(request=r, file=f)

            messages.success(request, "Cererea publică a fost actualizată (și fișierele au fost încărcate).")
        else:
            messages.success(request, "Cererea publică a fost actualizată.")

        return redirect("public_request_edit", pk=r.pk)

    return render(request, "portal/public_request_edit.html", {
        "req": r,
        "statuses": RequestStatus.objects.filter(is_active=True).order_by("order", "name"),
        "technicians": Technician.objects.filter(is_active=True).order_by("name"),
    })


@role_required(User.Role.ADMIN, User.Role.MANAGER)
@require_POST
def public_request_attachment_delete(request, pk):
    attachment = get_object_or_404(PublicRequestAttachment, pk=pk)
    req_id = attachment.request_id
    attachment.file.delete(save=False)
    attachment.delete()
    messages.success(request, "Documentul a fost șters.")
    return redirect("public_request_edit", pk=req_id)


@role_required(User.Role.ADMIN, User.Role.MANAGER)
@require_POST
def public_request_delete(request, pk):
    r = get_object_or_404(PublicRequest, pk=pk)
    r.delete()
    messages.success(request, "Cererea publică a fost ștearsă.")
    return redirect("portal_dashboard")


# ============================================================
# Dashboard core (build list) – REUTILIZAT + EXPORT
# ============================================================

def _get_dashboard_items(request) -> Dict[str, Any]:
    role = request.user.role

    if role == User.Role.CLIENT:
        tickets_qs = Ticket.objects.filter(created_by=request.user).select_related("created_by", "assigned_to")
        public_qs = PublicRequest.objects.filter(user=request.user).select_related("user", "status", "assigned_to")
    elif role == User.Role.TEHNICIAN:
        tickets_qs = Ticket.objects.filter(assigned_to=request.user).select_related("created_by", "assigned_to")
        public_qs = PublicRequest.objects.all().select_related("user", "status", "assigned_to")
    else:
        tickets_qs = Ticket.objects.all().select_related("created_by", "assigned_to")
        public_qs = PublicRequest.objects.all().select_related("user", "status", "assigned_to")

    show_public = request.GET.get("public", "1") == "1"
    show_client = request.GET.get("client", "1") == "1"
    show_intern = request.GET.get("intern", "1") == "1"

    q = _normalize_lower(request.GET.get("q"))
    f_type = _normalize(request.GET.get("type")).upper()
    f_nr = _normalize_lower(request.GET.get("nr"))
    f_name = _normalize_lower(request.GET.get("name"))
    f_phone = _normalize_lower(request.GET.get("phone"))
    f_status = _normalize_lower(request.GET.get("status"))
    f_assigned = _normalize_lower(request.GET.get("assigned"))

    quick = _normalize(request.GET.get("quick"))
    date_from = _parse_date(request.GET.get("date_from"))
    date_to = _parse_date(request.GET.get("date_to"))

    today = timezone.localdate()

    if quick == "today":
        date_from = today
        date_to = today
    elif quick == "yesterday":
        y = today - timedelta(days=1)
        date_from = y
        date_to = y
    elif quick == "7d":
        date_from = today - timedelta(days=6)
        date_to = today
    elif quick == "30d":
        date_from = today - timedelta(days=29)
        date_to = today
    elif quick == "this_month":
        date_from, date_to = _month_range(today)
    elif quick == "last_month":
        first_this, _ = _month_range(today)
        last_month_day = first_this - timedelta(days=1)
        date_from, date_to = _month_range(last_month_day)
    else:
        quick = ""

    dt_from = timezone.make_aware(datetime.combine(date_from, time.min)) if date_from else None
    dt_to = timezone.make_aware(datetime.combine(date_to, time.max)) if date_to else None

    if f_type in {"PUBLIC", "CLIENT", "INTERN"}:
        show_public = (f_type == "PUBLIC")
        show_client = (f_type == "CLIENT")
        show_intern = (f_type == "INTERN")

    default_public_status = RequestStatus.objects.filter(name__iexact="Neprocesat").first()

    items: List[Dict[str, Any]] = []

    if show_client or show_intern:
        for t in tickets_qs:
            source = _ticket_source(t)
            if (source == "CLIENT" and not show_client) or (source == "INTERN" and not show_intern):
                continue

            items.append({
                "kind": "ticket",
                "source": source,
                "nr": f"T-{t.id}",
                "client_name": _user_name(t.created_by),
                "client_phone": _user_phone(t.created_by),
                "status_name": t.get_status_display(),
                "assigned_name": (t.assigned_to.email if t.assigned_to else ""),
                "created_at": t.created_at,
                "pk": t.pk,
            })

    if show_public or show_client or show_intern:
        for r in public_qs:
            source = _public_source(r)
            if (source == "PUBLIC" and not show_public) or (source == "CLIENT" and not show_client) or (source == "INTERN" and not show_intern):
                continue

            assigned_name = ""
            if getattr(r, "assigned_to_id", None):
                assigned_name = getattr(r.assigned_to, "name", "") or ""

            items.append({
                "kind": "public",
                "source": source,
                "nr": f"P-{r.id}",
                "client_name": r.company or r.email,
                "client_phone": r.phone or "-",
                "status_name": (r.status.name if r.status else (default_public_status.name if default_public_status else "Neprocesat")),
                "assigned_name": assigned_name,
                "created_at": r.created_at,
                "pk": r.pk,
            })

    items_for_suggest = sorted(items, key=lambda x: x.get("created_at"), reverse=True)
    nr_suggestions = _build_autocomplete(items_for_suggest, "nr", 80)
    name_suggestions = _build_autocomplete(items_for_suggest, "client_name", 80)
    phone_suggestions = _build_autocomplete(items_for_suggest, "client_phone", 80)

    def match(it: Dict[str, Any]) -> bool:
        if f_type in {"PUBLIC", "CLIENT", "INTERN"} and it.get("source") != f_type:
            return False

        if q:
            hay = " ".join([
                str(it.get("source", "")),
                str(it.get("nr", "")),
                str(it.get("client_name", "")),
                str(it.get("client_phone", "")),
                str(it.get("status_name", "")),
                str(it.get("assigned_name", "")),
            ]).lower()
            if q not in hay:
                return False

        if f_nr and f_nr not in str(it.get("nr", "")).lower():
            return False
        if f_name and f_name not in str(it.get("client_name", "")).lower():
            return False
        if f_phone and f_phone not in str(it.get("client_phone", "")).lower():
            return False
        if f_status and f_status not in str(it.get("status_name", "")).lower():
            return False
        if f_assigned and f_assigned not in str(it.get("assigned_name", "")).lower():
            return False

        created = it.get("created_at")
        if dt_from and created and created < dt_from:
            return False
        if dt_to and created and created > dt_to:
            return False

        return True

    items = [it for it in items if match(it)]

    sort = _normalize(request.GET.get("sort")) or "created_at"
    direction = _normalize(request.GET.get("dir")) or "desc"
    if direction not in {"asc", "desc"}:
        direction = "desc"
    reverse = (direction == "desc")

    allowed = {"created_at", "nr", "client_name", "client_phone", "status_name", "assigned_name", "source"}
    if sort not in allowed:
        sort = "created_at"

    def sort_key(x: Dict[str, Any]):
        v = x.get(sort)
        return "" if v is None else v

    items.sort(key=sort_key, reverse=reverse)

    return {
        "items": items,
        "filter_public": show_public,
        "filter_client": show_client,
        "filter_intern": show_intern,

        "q": _normalize(request.GET.get("q")),
        "f_type": f_type,
        "f_nr": _normalize(request.GET.get("nr")),
        "f_name": _normalize(request.GET.get("name")),
        "f_phone": _normalize(request.GET.get("phone")),
        "f_status": _normalize(request.GET.get("status")),
        "f_assigned": _normalize(request.GET.get("assigned")),

        "date_from": date_from.isoformat() if date_from else "",
        "date_to": date_to.isoformat() if date_to else "",
        "quick": quick,

        "sort": sort,
        "dir": direction,

        "nr_suggestions": nr_suggestions,
        "name_suggestions": name_suggestions,
        "phone_suggestions": phone_suggestions,
    }


# ============================================================
# Dashboard view (HTML)
# ============================================================

@login_required
def dashboard(request):
    data = _get_dashboard_items(request)
    items = data["items"]

    per_page_choices = [25, 50, 100, 300, 500, 1000, 2000]
    per_page = _safe_int(request.GET.get("per_page"), 25)
    if per_page not in per_page_choices:
        per_page = 25

    paginator = Paginator(items, per_page)
    page_obj = paginator.get_page(_safe_int(request.GET.get("page"), 1))

    role = request.user.role
    statuses = RequestStatus.objects.all() if role in [User.Role.ADMIN, User.Role.MANAGER] else None
    technicians = Technician.objects.all() if role in [User.Role.ADMIN, User.Role.MANAGER] else None

    keep_qs = _qs_keep(request, drop=["page"])

    return render(request, "portal/dashboard.html", {
        "page_obj": page_obj,
        "per_page": per_page,
        "per_page_choices": per_page_choices,
        "keep_qs": keep_qs,

        "statuses": statuses,
        "technicians": technicians,
        "status_form": RequestStatusForm(),
        "technician_form": TechnicianForm(),

        **data,
    })


# ============================================================
# Export Excel (.xlsx)
# ============================================================

@login_required
def portal_export_xlsx(request):
    data = _get_dashboard_items(request)
    items = data["items"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Cereri"

    headers = ["Tip", "Nr cerere", "Nume client", "Telefon", "Status", "Alocat", "Creat", "Kind", "ID"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for it in items:
        created = it.get("created_at")
        if created and timezone.is_aware(created):
            created = timezone.localtime(created).replace(tzinfo=None)

        ws.append([
            it.get("source", ""),
            it.get("nr", ""),
            it.get("client_name", ""),
            it.get("client_phone", ""),
            it.get("status_name", ""),
            it.get("assigned_name", ""),
            created,
            it.get("kind", ""),
            it.get("pk", ""),
        ])

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=7).number_format = "yyyy-mm-dd hh:mm"

    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 50)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"portal_export_{timezone.localdate().isoformat()}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ============================================================
# Ticket create / assign / update status
# ============================================================

@role_required(User.Role.CLIENT)
def ticket_create(request):
    ip = get_client_ip(request)

    # ✅ DB rate limit: max 10 POST/minut per user (sau IP)
    if request.method == "POST":
        if db_rate_limit_hit(request, "rl_ticket_create_60s", limit=10, window=timedelta(seconds=60)):
            log_abuse(request, "rl_ticket_create_block")
            maybe_block_ip(ip, "rl_ticket_create_block")
            return HttpResponse("Prea multe cereri într-un timp scurt. Reîncearcă mai târziu.", status=429)
        db_rate_limit_mark(request, "rl_ticket_create_60s")

    # ✅ limită business: max 3 cereri/oră / user
    one_hour_ago = timezone.now() - timedelta(hours=1)
    if Ticket.objects.filter(created_by=request.user, created_at__gte=one_hour_ago).count() >= 3:
        log_abuse(request, "limit_3_per_hour")
        maybe_block_ip(ip, "limit_3_per_hour")
        return HttpResponse("Ai depășit limita de 3 cereri/oră.", status=429)

    if request.method == "POST":
        form = TicketCreateForm(request.POST)
        if form.is_valid():
            t = form.save(commit=False)
            t.created_by = request.user

            if hasattr(t, "ip_address"):
                t.ip_address = ip

            t.save()
            messages.success(request, "Cererea a fost trimisă.")
            return redirect("portal_dashboard")
        else:
            log_abuse(request, "ticket_form_invalid")
            maybe_block_ip(ip, "ticket_form_invalid")
    else:
        form = TicketCreateForm()

    return render(request, "portal/ticket_create.html", {"form": form})


@role_required(User.Role.ADMIN, User.Role.MANAGER)
def ticket_assign(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    technicians = User.objects.filter(role=User.Role.TEHNICIAN, is_active=True)

    if request.method == "POST":
        tech_id = request.POST.get("assigned_to") or None
        status = request.POST.get("status")

        ticket.assigned_to_id = int(tech_id) if tech_id else None
        if status and status in dict(Ticket.Status.choices):
            ticket.status = status
        ticket.save()

        messages.success(request, "Ticket actualizat.")
        return redirect("portal_dashboard")

    return render(request, "portal/ticket_assign.html", {
        "ticket": ticket,
        "technicians": technicians,
        "statuses": Ticket.Status.choices,
    })


@role_required(User.Role.ADMIN, User.Role.MANAGER, User.Role.TEHNICIAN)
def ticket_update_status(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)

    if request.user.role == User.Role.TEHNICIAN and ticket.assigned_to_id != request.user.id:
        return render(request, "portal/forbidden.html", status=403)

    if request.method == "POST":
        status = request.POST.get("status")
        if status in dict(Ticket.Status.choices):
            ticket.status = status
            ticket.save()
            messages.success(request, "Status actualizat.")
        return redirect("portal_dashboard")

    return render(request, "portal/ticket_update_status.html", {
        "ticket": ticket,
        "statuses": Ticket.Status.choices
    })


# ============================================================
# Site settings
# ============================================================

@role_required(User.Role.ADMIN, User.Role.MANAGER)
def site_settings_view(request):
    settings_obj = get_site_settings()

    if request.method == "POST":
        form = SitePagesSettingsForm(request.POST, instance=settings_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Setările au fost salvate.")
            return redirect("portal_site_settings")
    else:
        form = SitePagesSettingsForm(instance=settings_obj)

    return render(request, "portal/site_settings.html", {"form": form})


# ============================================================
# Inline create (dashboard tabs)
# ============================================================

@role_required(User.Role.ADMIN, User.Role.MANAGER)
def status_create_inline(request):
    if request.method == "POST":
        form = RequestStatusForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Status creat.")
    return redirect("portal_dashboard")


@role_required(User.Role.ADMIN, User.Role.MANAGER)
def technician_create_inline(request):
    if request.method == "POST":
        form = TechnicianForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Tehnician adăugat.")
    return redirect("portal_dashboard")


# ============================================================
# STATUSURI CRUD
# ============================================================

@role_required(User.Role.ADMIN, User.Role.MANAGER)
def status_list(request):
    statuses = RequestStatus.objects.all()
    return render(request, "portal/status_list.html", {"statuses": statuses})


@role_required(User.Role.ADMIN, User.Role.MANAGER)
def status_create(request):
    if request.method == "POST":
        form = RequestStatusForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Status creat.")
            return redirect("status_list")
    else:
        form = RequestStatusForm()
    return render(request, "portal/status_form.html", {"form": form, "mode": "create"})


@role_required(User.Role.ADMIN, User.Role.MANAGER)
def status_update(request, pk):
    obj = get_object_or_404(RequestStatus, pk=pk)
    if request.method == "POST":
        form = RequestStatusForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Status actualizat.")
            return redirect("status_list")
    else:
        form = RequestStatusForm(instance=obj)
    return render(request, "portal/status_form.html", {"form": form, "mode": "edit", "obj": obj})


@role_required(User.Role.ADMIN, User.Role.MANAGER)
def status_delete(request, pk):
    obj = get_object_or_404(RequestStatus, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Status șters.")
        return redirect("status_list")
    return render(request, "portal/status_confirm_delete.html", {"obj": obj})


# ============================================================
# TEHNICIENI CRUD
# ============================================================

@role_required(User.Role.ADMIN, User.Role.MANAGER)
def tech_list(request):
    techs = Technician.objects.all()
    return render(request, "portal/tech_list.html", {"techs": techs})


@role_required(User.Role.ADMIN, User.Role.MANAGER)
def tech_create(request):
    if request.method == "POST":
        form = TechnicianForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Tehnician adăugat.")
            return redirect("tech_list")
    else:
        form = TechnicianForm()
    return render(request, "portal/tech_form.html", {"form": form, "mode": "create"})


@role_required(User.Role.ADMIN, User.Role.MANAGER)
def tech_update(request, pk):
    obj = get_object_or_404(Technician, pk=pk)
    if request.method == "POST":
        form = TechnicianForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Tehnician actualizat.")
            return redirect("tech_list")
    else:
        form = TechnicianForm(instance=obj)
    return render(request, "portal/tech_form.html", {"form": form, "mode": "edit", "obj": obj})


@role_required(User.Role.ADMIN, User.Role.MANAGER)
def tech_delete(request, pk):
    obj = get_object_or_404(Technician, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Tehnician șters.")
        return redirect("tech_list")
    return render(request, "portal/tech_confirm_delete.html", {"obj": obj})
